# -*- coding: utf-8 -*-
# @Time    : 2020/3/18 15:36
# @Author  : chengrun.niu
# @Email   : chengrun.niu@aispeech.com
import requests
import uuid
import json
from openpyxl import load_workbook
import time
from openpyxl.styles import Font
from openpyxl.styles.colors import RED,BLACK
from openpyxl.styles import NamedStyle
class work():

    def __init__(self):
        #self.url = "http://dm-runtime-alpha.talkinggenie.com/callcenter/nlu"    #alpha
        self.url = "http://dm-runtime.talkinggenie.com/callcenter/nlu"    #线上
        self.path = r"C:\Users\wwh05\Documents\工作资料\python\python-test\SAAS\沈阳银信盛京M1催收\沈阳银信盛京M1催收用例.xlsx"
        self.productid = "914013350"
        self.sheet = "Sheet3"
       # self.Robot = "S6681935000923500544"

    def url_1(self,senderid,query):
        body = {
            "senderId": senderid,
            "ses1sionId": senderid,
            "asrConfidence": 1,
            "productId": self.productid,
           # "robotId":self.Robot,
            "query": query,
            "message": {}
                    }
        # 请求url
        response = requests.post(url=self.url, json=body)
        #response = requests.post(url=self.url, data=json.dumps(body))


        time.sleep(0.2)
        """
        print(response.encoding)
        print(response.content.decode())
        print(type(response.content.decode()))
        response.content得到的是一个bytes
        response.content.decode()中decode将bytes转换成字符串
        return response.content.decode()
        response.text得到的是字符串
        """
        return response.text

    def run(self):
        # 获取一个execl文件，赋值给一个变量
        wb = load_workbook(self.path)
        # 获取execl中表格名称，赋值给一个对象
        sheet = wb[self.sheet]
        # 已列表的形式循环遍历，并按照列获取值 从1索引到最后 index 索引 val 获取的execl整列的值
        #for index,val in enumerate(list(sheet.columns)[1][1:]):
        for val in list(sheet.columns)[1][1:]:
            if val.value is None:
                pass
            else:
                if val.value == "signal=newCall":
                    senderid = str(uuid.uuid4())
                    result_sendid = 'F' + str(val.row)
                    sheet[result_sendid] = senderid
                # if val.value == "signal=outbound":
                #print(senderid)
               # print(val.column)
                #print(val.row)
                result = self.url_1(senderid, val.value)
                #print(result)
                #print(type(result))
                #print(val.value)
                # elif val.value == "falsee":
                #     result = self.url_2(senderid,val.value,"1")
                #     print(val.value)
                # elif val.value == "hungup":
                #     result = self.url_2(senderid, val.value,"2")
                #     print(val.value)
                # else:
                #     result = self.url_2(senderid, val.value,"10")
                #     print(val.value)
                #print(str(val.row))
                try:

                    result_excelid = 'C' + str(val.row)
                    #print(result)
                    #a = json.loads(result)
                    #print(a)
                    sheet[result_excelid] = json.loads(result)["voiceText"].strip()
                    #sheet[result_excelid] = json.loads(result)["best"]["reply"]["voiceText"].strip()

                except:
                    result_excelid = 'C' + str(val.row)
                    sheet[result_excelid] = json.loads(result)["best"]["reply"]["voiceText"].strip()
                    a= sheet[result_excelid]
                    print(val.value)
                    print(a.value)

                try:
                    topicCode = json.loads(result)["best"]["topicCode"]
                except:
                    topicCode = "null"
                result_excelid = 'G' + str(val.row)
                sheet[result_excelid] = topicCode



                if topicCode == "1005":

                    body = {"senderId": senderid, "sessionId": senderid, "asrConfidence": 1, "productId": self.productid,
                            "query": "signal=hangUp",
                            "message":
                                {}}
                    # time.sleep(0.02)
                    re = requests.post(url=self.url, json=body)


                    try:
                        message = json.loads(re.text)["best"]["reply"]["message"]
                        re_message = 'H' + str(val.row)
                        sheet[re_message] = str(message)
                    except:
                        pass
                elif topicCode == "1002":
                    message = json.loads(result)["best"]["reply"]["message"]
                    print("转人工回传参数： ")

                    re_message = 'H' + str(val.row)
                    sheet[re_message] = str(message)
                    # print(message)

                else:
                    pass
        wb.save(self.path)

    def compare_query(self):
        wb = load_workbook(self.path)
        sheet = wb[self.sheet]
        font_r = Font(color=RED)
        font_b = Font(color=BLACK)
        for id, val in enumerate(list(sheet.columns)[2][1:]):
            #print(id, val)
            shiji = 'C' + str(val.row)
            #print(shiji)
            yuqi = 'D' + str(val.row)
            yuqi_result = sheet[yuqi]
            #print(yuqi_result)
            #print(yuqi_result.value)
            jieguo = 'E' + str(val.row)

            if yuqi_result.value is None and val.value is None:
                sheet[jieguo] = '成功'
                sheet[jieguo].font = font_b
                sheet[shiji].font = font_b
                sheet[yuqi].font = font_b
            if yuqi_result.value is None or val.value is None:
                sheet[jieguo] = '失败'
                sheet[jieguo].font = font_r
                sheet[shiji].font = font_r
                sheet[yuqi].font = font_r
            elif yuqi_result.value.strip() == val.value.strip():
                sheet[jieguo] = '通过'
                sheet[jieguo].font = font_b
                sheet[shiji].font = font_b
                sheet[yuqi].font = font_b
            else:
                sheet[jieguo] = '失败'
                sheet[jieguo].font = font_r
                sheet[shiji].font = font_r
                sheet[yuqi].font = font_r
        wb.save(self.path)



if __name__ == '__main__':
    sf = work()
    #shee = "Sheet" + str(6)
    sf.run()
    sf.compare_query()
