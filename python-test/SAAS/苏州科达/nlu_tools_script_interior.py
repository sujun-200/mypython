#!/usr/bin/env python
# coding=utf-8

## ----------------------------------------
import sys
import json

import requests
import argparse, textwrap
import uuid
import openpyxl

URL_MOULD = "{}/alarm/event_extraction"

punctuation_str = '！？。；'


def extraction(args):
    file_name = args.file_name
    sheet_index = int(args.sheet_index) - 1
    row = int(args.row)
    column = int(args.column)
    pick_type = args.pick_type
    url = URL_MOULD.format(args.address)
    inverse_url = args.inverse_address
    pid = args.pid

    query_map, city_code_map = load_file(file_name, sheet_index, row, column, pick_type)

    if query_map is None or len(query_map) == 0:
        print('this file：{} not found data，please check file params'.format(file_name))
        return

    wb = openpyxl.load_workbook('./{}'.format(file_name))
    ws = wb[wb.sheetnames[sheet_index]]
    num = 0
    # ws.cell(row=(row-1), column=(column + 1)).value = '地址识别'
    # ws.cell(row=(row-1), column=(column + 2)).value = '地址补全'
    # ws.cell(row=(row-1), column=(column + 3)).value = '警情要素'
    for index1, index2 in zip(query_map, city_code_map):
        num += 1
        session_id = create_session_id()
        result_list = []
        for query in query_map[index1]:
            result_list.append(access(url, session_id, query, city_code_map[index2], pid))
        record = handler_result(num, result_list, inverse_url)
        info = ''
        raw_address = '{}'
        completed_address = '{}'
        for k, v in record.items():
            if '原文地址' in k:
                raw_address = raw_address.format(v)
            elif '警情地址' in k:
                completed_address = completed_address.format(v)
            else:
                info = info + '{}：{}\n'.format(k, v)
        ws.cell(row=row, column=(column + 1)).value = raw_address
        ws.cell(row=row, column=(column + 2)).value = completed_address
        ws.cell(row=row, column=(column + 3)).value = info
        row += 1
    wb.save('./{}'.format(file_name))

    print('执行完成！')


def load_file(file_name, sheet_index, row, column, pick_type):
    query_map = {}
    city_code_map = {}
    try:
        wb = openpyxl.load_workbook('./{}'.format(file_name))
        ws = wb[wb.sheetnames[sheet_index]]
        while True:
            queries = []
            queries_value = ws.cell(row=row, column=column).value
            city_code_value = ws.cell(row=row, column=(column - 2)).value

            if queries_value is None or len(queries_value) == 0:
                break
            queries_value = str(queries_value)
            city_code_value = str(city_code_value)
            if '换行' in pick_type:
                if '\n' in queries_value:
                    arr = queries_value.split('\n')
                    for line in arr:
                        if len(line) > 0:
                            queries.append(line)
                else:
                    queries.append(queries_value)
            elif '标点' in pick_type:
                last_index = 0
                for i in range(0, len(queries_value)):
                    if queries_value[i] in punctuation_str:
                        queries.append(queries_value[last_index:i])
                        last_index = i + 1
                        print(1)
                if last_index < len(queries_value):
                    queries.append(queries_value[last_index:len(queries_value)])
            if len(queries) > 0:
                query_map[row] = queries
                city_code_map[row] = city_code_value
            row += 1
    except:
        print('occur error while load case where file name is'
              '：{},\tsheet name is：{},\trow is：{},\tcolumn is：{},\tpick_type is'
              '：{}'.format(file_name, sheet_index, row, column, pick_type))
    return query_map, city_code_map


def access(url, session_id, query, city_code, pid):
    result = ''
    if city_code is None or len(city_code) <= 0:
        city_code = ''
    else:
        city_code = '0{}'.format(city_code)
    query = query.replace('！', '').replace('？', '').replace('。', '').replace('？', '；')
    try:
        payload = {
            'query': query,
            'sessionId': session_id,
            'cityCode': city_code,
            'dialog': {
                'productId': pid
            }
        }
        header = {
            'Authorization': '98926d7091bf4de28002b2eba7526a0e',
            'Content-Type': 'application/json',
            'sign': 'ba1dbdb888031231a675e1657117dd8a'
        }
        resp = requests.post(url, data=json.dumps(payload), headers=header)
        result = resp.text
    except:
        print('occur error where url is：{} and query is: {}'.format(url, query))
    return result


def inverse(query, url):
    result = query
    try:
        header = {
            'Content-Type': 'application/json'
        }
        payload = [query]
        resp = requests.post(url, data=json.dumps(payload), headers=header)
        text = json.loads(resp.text)
        context = text['data']['result'][0]
        if len(context) is not None and len(context) > 0:
            result = context
    except:
        print('逆文本调用出错，请检查地址，url是：{}'.format(url))
    return result


def create_session_id():
    result = str(uuid.uuid1())
    return result


def handler_result(num, result_list, inverse_url):
    if result_list is None or len(result_list) == 0:
        print('the case {} result is empty'.format(num))
    record = {}
    try:
        for item in result_list:
            item_json = json.loads(item)
            result_pick(record, item_json, 'raw_address', '原文地址', inverse_url)
            result_pick(record, item_json, 'completed_address', '警情地址', inverse_url)
            result_pick(record, item_json, 'police_type', '灾害类型', inverse_url)
            result_pick(record, item_json, 'floor', '燃烧楼层', inverse_url)
            result_pick(record, item_json, 'is_trapped', '人员被困情况', inverse_url)
            result_pick(record, item_json, 'fire_place', '火灾场所', inverse_url)
            result_pick(record, item_json, 'fire_object', '燃烧对象', inverse_url)
            result_pick(record, item_json, 'structure', '建筑结构', inverse_url)
            result_pick(record, item_json, 'smog_condition', '烟雾状况', inverse_url)
    except:
        print('occur error where handler result for case：{} and result is: {}'.format(num, result_list))
    return record


def result_pick(record, item, key, format_key, inverse_url):
    pick = item.get(key, '')
    if len(pick) > 0:
        pick = inverse(pick, inverse_url)

    if '原文地址' in format_key or '警情地址' in format_key:
        if len(pick) > 0 and (record.get(format_key) is None or len(record.get(format_key)) <= 0):
            record[format_key] = pick + ';\n'
        elif len(pick) > 0 and record.get(format_key) is not None and len(record.get(format_key)) > 0:
            record[format_key] = record.get(format_key) + pick + ';\n'
        elif len(pick) <= 0 and (record.get(format_key) is None or len(record.get(format_key)) <= 0):
            record[format_key] = 'null;\n'
        elif len(pick) <= 0 and record.get(format_key) is not None and len(record.get(format_key)) > 0:
            record[format_key] = record.get(format_key) + 'null;\n'
    else:
        if len(pick) > 0:
            record[format_key] = pick
        elif record.get(format_key) is None or len(record.get(format_key)) <= 0:
            record[format_key] = pick


def argparser():
    """
    解析脚本参数
    """
    parser = argparse.ArgumentParser()

    parser.add_argument("-address", help="ip和端口", type=str, default='https://api-alpha.talkinggenie.com/ccdmctrl')

    parser.add_argument("-active", type=str, default='extraction')

    parser.add_argument("-file_name", help="测试例子的文件名字", type=str, default='test.xlsx')

    parser.add_argument("-sheet_index", help="测试文件的sheet下标，从1开始计算", type=str, default=1)

    parser.add_argument("-row", help="测试文件的例子从第几行开始", type=str, default=2)

    parser.add_argument("-column", help="测试文件的例子从第几列开始", type=str, default=4)

    parser.add_argument("-pick_type", help="句子挑拣规则，换行、标点", type=str, default='换行')

    parser.add_argument("-pid", help="产品id", type=str, default='914002890')

    parser.add_argument("-inverse_address", help="逆文本地址", type=str, default='http://ezmt.duiopen.com/ezapi/inverse  ')

    args = parser.parse_args()

    return args


def default(args):
    print("\n没有定义相应的动作，需要指定参数:  -file_name\n\n"
          "example:   -file_name=test.xlsx")


def main(argv):
    """

    """
    args = argparser()

    global URL_MOULD

    url = URL_MOULD.format(args.address)

    print("当前使用的执行地址是: {}".format(url))

    switch = {
        "extraction": extraction
    }

    active = args.active
    switch.get(active, default)(args)


if __name__ == "__main__":
    main(sys.argv)
